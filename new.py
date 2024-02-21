import pygame
import random
import math
import tkinter as tk
import tkinter as tk
from tkinter import font
from tkinter import ttk
from tkinter import messagebox
import time
import os
import openpyxl
from openpyxl.styles import Alignment
import datetime

pygame.mixer.init()
prey_move_sound = pygame.mixer.Sound("prey.mp3")
kill = pygame.mixer.Sound("kill.mp3")
time_up = pygame.mixer.SoundType("time.mp3")

pygame.init()
display_info = pygame.display.Info()

# Set screen dimensions
SCREEN_WIDTH = display_info.current_w
SCREEN_HEIGHT = display_info.current_h

# Define colors
WHITE = (255, 255, 255)
GREEN = (0, 255, 0)
RED = (255, 0, 0)
BLACK = (0, 0, 0)

# Set parameters for prey and predators
PREY_REPRODUCTION_RATE = 0.01
PREDATOR_REPRODUCTION_RATE = 0.01
PREDATOR_MORTALITY_RATE = 0
MOVEMENT_SPEED = 5
SIGHT_RANGE = 25
PREDATOR_SIZE = .40
CHANGE_DIRECTION_INTERVAL = 60
remaining_prey = 0

class Prey:
    def __init__(self):
        self.x = random.randint(0, SCREEN_WIDTH)
        self.y = random.randint(0, SCREEN_HEIGHT)
        self.direction = random.uniform(0, 2*math.pi) 
        self.alive = True
        
    def move(self, predators, all_prey):
        speed = STEP_SIZE

        # Check if prey is being chased by any predator
        for predator in predators:
            distance = math.sqrt((self.x - predator.x)**2 + (self.y - predator.y)**2)
            if distance < NOISE_RADIUS:
                away_angle = math.atan2(self.y - predator.y, self.x - predator.x)
                self.direction = away_angle + 30
                speed = RUN_SPEED
                break

        # Avoid clustering with other predators
        for prey in all_prey:
            if prey != self:
                distance = math.sqrt((self.x - prey.x)**2 + (self.y - prey.y)**2)
                if distance < 10:
                    away_angle = math.atan2(self.y - prey.y, self.x - prey.x)
                    self.direction = away_angle + 30
                    speed = STEP_SIZE
                    break

        # Move the prey
        self.x += speed * math.cos(self.direction)
        self.y += speed * math.sin(self.direction)

        # Wrap around screen edges
        self.x %= SCREEN_WIDTH
        self.y %= SCREEN_HEIGHT

    def collide(self, other):
        distance = math.sqrt((self.x - other.x)**2 + (self.y - other.y)**2)
        if distance < 10:  
            collision_angle = math.atan2(other.y - self.y, other.x - self.x)
            
            # Reflect directions
            self.direction = 2 * collision_angle - self.direction
            other.direction = 2 * collision_angle - other.direction

    def reproduce(self):
        if random.random() < PREY_REPRODUCTION_RATE:
            return Prey()
        else:
            return None

    def draw(self, screen):
        pygame.draw.circle(screen, GREEN, (int(self.x), int(self.y)), 5)
        end_x = self.x + 10 * math.cos(self.direction)
        end_y = self.y + 10 * math.sin(self.direction)
        pygame.draw.line(screen, WHITE, (self.x, self.y), (end_x, end_y))

class Predator:
    def __init__(self):
        self.x = random.randint(0, SCREEN_WIDTH)
        self.y = random.randint(0, SCREEN_HEIGHT)
        self.alive = True
        self.direction = random.uniform(0, 2*math.pi)
        self.change_direction_counter = 0
        self.detected_prey = None
        self.angle = 0
        
    def move(self, prey_population, predator_population):
        # Rotate the predator
        self.angle += 0.05  

        # Check if any prey is within the flee distance and in front of the predator
        prey_within_flee_distance = False
        for prey in prey_population:
            distance_to_prey = self.detect_distance(prey)
            angle_to_prey = math.atan2(prey.y - self.y, prey.x - self.x)
            angle_difference = abs(angle_to_prey - self.direction)
            if distance_to_prey < FLEE_DISTANCE and angle_difference < math.pi/3:
                prey_within_flee_distance = True
                break

        if prey_within_flee_distance:
            closest_prey = self.detect_closest_prey(prey_population)
            prey_direction = math.atan2(closest_prey.y - self.y, closest_prey.x - self.x)
            self.direction = prey_direction

            # Increase speed when chasing prey
            self.speed = CHASE_SPEED
        else:
            # Move randomly if no prey nearby
            self.direction += random.uniform(-0.1, 0.1)

            # Set regular speed when not chasing prey
            self.speed = REGULAR_SPEED

        # Avoid clustering with other predators
        for predator in predator_population:
            if predator != self:
                distance = math.sqrt((self.x - predator.x)**2 + (self.y - predator.y)**2)
                if distance < 20:
                    # Calculate angle away from nearby predator
                    away_angle = math.atan2(self.y - predator.y, self.x - predator.x)
                    self.direction = away_angle + random.uniform(-0.1, 0.1)
                    break

        # Update position based on direction and speed
        self.x += self.speed * math.cos(self.direction)
        self.y += self.speed * math.sin(self.direction)

        # Wrap around screen edges
        self.x %= SCREEN_WIDTH
        self.y %= SCREEN_HEIGHT

    # detect the closest prey
    def detect_closest_prey(self, prey_population):
        closest_prey = None
        closest_distance = float('inf')
        for prey in prey_population:
            distance = self.detect_distance(prey)
            if distance < closest_distance:
                closest_distance = distance
                closest_prey = prey
        return closest_prey

    # calculate distance to prey
    def detect_distance(self, prey):
        return math.sqrt((self.x - prey.x)**2 + (self.y - prey.y)**2)

    def die(self):
        if random.random() < PREDATOR_MORTALITY_RATE:
            self.alive = False

    def draw(self, screen):
        pygame.draw.circle(screen, RED, (int(self.x), int(self.y)), 7.5)

        # Draw circular border for flee radius
        # pygame.draw.circle(screen, WHITE, (int(self.x), int(self.y)), FLEE_DISTANCE, 2)

        end_x = self.x + 10 * math.cos(self.direction)
        end_y = self.y + 10 * math.sin(self.direction)
        pygame.draw.line(screen, WHITE, (self.x, self.y), (end_x, end_y))

        

def run_simulation(num_prey, prey_speed, num_predators, predator_speed, avoid_speed, noise_radius, chase_speed, chase_radius):
    global remaining_prey
    global STEP_SIZE
    global REGULAR_SPEED
    global RUN_SPEED
    global NOISE_RADIUS
    global CHASE_SPEED
    global FLEE_DISTANCE

    STEP_SIZE = prey_speed
    REGULAR_SPEED = predator_speed
    RUN_SPEED = avoid_speed
    NOISE_RADIUS = noise_radius
    CHASE_SPEED = chase_speed
    FLEE_DISTANCE = chase_radius
    
    pygame.init()
    screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))
    pygame.display.set_caption("Predator and Prey Simulation")

    # Create prey and predator populations
    prey_population = [Prey() for _ in range(num_prey)]
    predator_population = [Predator() for _ in range(num_predators)]

    # Main loop
    clock = pygame.time.Clock()
    running = True

    # Stopwatch variables
    start_time = time.time()
    stopwatch_font = pygame.font.Font(None, 25)

    # Create back button
    back_button = pygame.Rect(20, 20, 60, 35)
    font = pygame.font.Font(None, 25)
    back_text = font.render("Stop", True, WHITE)

    while running:
        screen.fill(BLACK)

        # Event handling
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            elif event.type == pygame.MOUSEBUTTONDOWN:
                if event.button == 1:
                    if back_button.collidepoint(event.pos): 
                        pygame.quit()
                        return get_user_input()

        # Move prey
        for prey in prey_population:
            prey.move(predator_population, prey_population)  
            prey.draw(screen)

        # Move predators and handle interactions with prey
        for predator in predator_population:
            predator.move(prey_population, predator_population)  
            predator.draw(screen)
            prey_to_remove = None
            for prey in prey_population:
                # Check for collision between predator and prey
                distance = math.sqrt((predator.x - prey.x)**2 + (predator.y - prey.y)**2)
                if distance < SIGHT_RANGE: 
                    prey_to_remove = prey
                    predator.die()
                    remaining_prey += 1
                    kill.play()
                    break 
            if prey_to_remove:
                prey_population.remove(prey_to_remove)

        # Remove dead predators
        predator_population = [predator for predator in predator_population if predator.alive]

        # Display remaining prey count and time at the bottom
        remaining_prey_count = len(prey_population)

        time_text_surface = font.render(f"Prey Remaining: {remaining_prey_count}", True, WHITE)
        time_text_rect = time_text_surface.get_rect(midbottom=(SCREEN_WIDTH // 2, SCREEN_HEIGHT))
        screen.blit(time_text_surface, time_text_rect)

        pygame.draw.rect(screen, GREEN, back_button)
        screen.blit(back_text, (back_button.x + 10, back_button.y + 10))

        # Stopwatch
        elapsed_time = time.time() - start_time
        minutes = int(elapsed_time // 60)
        seconds = int(elapsed_time % 60)    
        stopwatch_text = f"{minutes:02d}:{seconds:02d}"
        stopwatch_surface = stopwatch_font.render(stopwatch_text, True, WHITE)
        stopwatch_rect = stopwatch_surface.get_rect(topright=(SCREEN_WIDTH - 20, 20))
        screen.blit(stopwatch_surface, stopwatch_rect)

        if remaining_prey_count == 0:
            time_up.play()
            pygame.quit()
            
            results(num_prey, num_predators, noise_radius, FLEE_DISTANCE, prey_speed, predator_speed, stopwatch_text)

        # Update display
        pygame.display.flip()
        clock.tick(60)

    # Quit pygame
    pygame.quit()


def results(prey_population, predator_population, prey_sight_range, predator_sight_range, prey_speed, predator_speed, total_catch_time):
    current_date_time = datetime.datetime.now().strftime("%B %d, %Y %I:%M %p")

    input_window = tk.Toplevel()
    input_window.title("Results")

    def see_results():
        os.system("start player.xlsx")  
    see_results_button = tk.Button(input_window, text="See Results", font=font.Font(size=12), command=see_results)
    see_results_button.grid(row=1, column=0, pady=10, padx=10)

    wb = openpyxl.load_workbook("player.xlsx")
    sheet = wb.active

    next_row = sheet.max_row + 1

    sheet.cell(row=next_row, column=1).value = current_date_time
    sheet.cell(row=next_row, column=2).value = prey_population
    sheet.cell(row=next_row, column=3).value = predator_population
    sheet.cell(row=next_row, column=4).value = prey_sight_range
    sheet.cell(row=next_row, column=5).value = predator_sight_range
    sheet.cell(row=next_row, column=6).value = prey_speed
    sheet.cell(row=next_row, column=7).value = predator_speed
    sheet.cell(row=next_row, column=8).value = total_catch_time

    for col in range(1, 9):
        sheet.cell(row=next_row, column=col).alignment = Alignment(horizontal='center')

    # Save the workbook
    wb.save("player.xlsx")

    tk.Label(input_window, text=total_catch_time, font=font.Font(size=18)).grid(row=0, column=0, columnspan=2, padx=110, pady=20)

    okay_button = tk.Button(input_window, text="Play Again", font=font.Font(size=12), command=lambda: [input_window.destroy(), get_user_input()])
    okay_button.grid(row=1, column=1, pady=10, padx=0)

    window_width = 300 
    window_height = 150
    screen_width = input_window.winfo_screenwidth()
    screen_height = input_window.winfo_screenheight()
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2
    input_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    input_window.mainloop()


def get_user_input():
    root = tk.Tk()
    root.withdraw()

    input_window = tk.Toplevel(root)
    input_window.title("Predator and Prey")

    # Define separator
    separator = ttk.Separator(input_window, orient='vertical')
    separator.grid(row=0, column=2, rowspan=5, sticky='ns', padx=20)

    tk.Label(input_window, text="Prey", font=font.Font(size=18)).grid(row=0, column=0, columnspan=2, pady=10)
    tk.Label(input_window, text="Predator", font=font.Font(size=18)).grid(row=0, column=3, columnspan=4, pady=10)

    separator.grid(row=0, column=2, rowspan=5, sticky='ns', padx=20)

    tk.Label(input_window, text="Prey", font=tk.font.Font(size=18)).grid(row=0, column=0, columnspan=2, pady=10)
    tk.Label(input_window, text="Predator", font=tk.font.Font(size=18)).grid(row=0, column=3, columnspan=4, pady=10)

    tk.Label(input_window, text="Population:", anchor="w", justify="left").grid(row=1, column=0, padx=(30,10), pady=5, sticky="w")
    entry_prey = tk.Entry(input_window, width=10)
    entry_prey.grid(row=1, column=1, padx=5, pady=5)

    tk.Label(input_window, text="Speed:", anchor="w", justify="left").grid(row=2, column=0, padx=(30,10), pady=5, sticky="w")
    entry_prey_speed = tk.Entry(input_window, width=10)
    entry_prey_speed.grid(row=2, column=1, padx=5, pady=5)
    entry_prey_speed.insert(0, "1.1")

    tk.Label(input_window, text="Avoid Speed:", anchor="w", justify="left").grid(row=3, column=0, padx=(30,10), pady=5, sticky="w")
    entry_prey_avoid_speed = tk.Entry(input_window, width=10)
    entry_prey_avoid_speed.grid(row=3, column=1, padx=5, pady=5)
    entry_prey_avoid_speed.insert(0, "1.1")

    tk.Label(input_window, text="Sight Range:", anchor="w", justify="left").grid(row=4, column=0, padx=(30,10), pady=5, sticky="w")
    noise_radius = tk.Entry(input_window, width=10)
    noise_radius.grid(row=4, column=1, padx=5, pady=5)
    noise_radius.insert(0, "360")

    tk.Label(input_window, text="Population:", anchor="w", justify="left").grid(row=1, column=3, pady=5, sticky="w")
    entry_predators = tk.Entry(input_window, width=10)
    entry_predators.grid(row=1, column=4, padx=5, pady=5)

    tk.Label(input_window, text="Speed:", anchor="w", justify="left").grid(row=2, column=3, pady=5, sticky="w")
    entry_predator_speed = tk.Entry(input_window, width=10)
    entry_predator_speed.grid(row=2, column=4, padx=5, pady=5)
    entry_predator_speed.insert(0, "1.1")

    tk.Label(input_window, text="Chase Speed:", anchor="w", justify="left").grid(row=3, column=3, pady=5, sticky="w")
    entry_predator_chase_speed = tk.Entry(input_window, width=10)
    entry_predator_chase_speed.grid(row=3, column=4, padx=5, pady=5)
    entry_predator_chase_speed.insert(0, "1.1")

    tk.Label(input_window, text="Sight Range:", anchor="w", justify="left").grid(row=4, column=3, pady=5, sticky="w")
    chase_redius = tk.Entry(input_window, width=10)
    chase_redius.grid(row=4, column=4, padx=5, pady=5)
    chase_redius.insert(0, "180")

    def submit():
        try:
            num_prey = int(entry_prey.get())
            prey_speed = float(entry_prey_speed.get())
            num_predators = int(entry_predators.get())
            predator_speed = float(entry_predator_speed.get())
            avoid_speed = float(entry_prey_avoid_speed.get())
            noise_radiuss = float(noise_radius.get())
            chase_speed = float(entry_predator_chase_speed.get())
            chase_radiuss = int(chase_redius.get())

            if num_prey <= 0:
                messagebox.showerror("Prey Error", "Number of prey must be at least 1!")
            elif prey_speed < 0:
                messagebox.showerror("Prey Speed Error", "Speed of prey must be greater than 0!")
            elif num_predators <= 0:
                messagebox.showerror("Predator Error", "Number of predators must be at least 1!")
            elif predator_speed < 0:
                messagebox.showerror("Predator Speed Error", "Speed of predators must be greater than 0!")
            elif avoid_speed < 0:
                messagebox.showerror("Prey Avoid Speed Error", "Avoid speed must greater than 0!")
            elif noise_radiuss < 0:
                messagebox.showerror("Prey Avoid Angle Error", "Avoid angle must greater than 0!")
            elif chase_speed < 0:
                messagebox.showerror("Predator Chase Speed Error", "Chase Speed must greater than 0!")
            elif chase_radiuss < 0:
                messagebox.showerror("Predator Chase Radius Error", "Chase Radius must greater than 0!")
            else:
                input_window.destroy()
                run_simulation(num_prey, prey_speed, num_predators, predator_speed, avoid_speed, noise_radiuss, chase_speed, chase_radiuss)
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numbers.")

    button_submit = tk.Button(input_window, text="Submit", command=submit)
    button_submit.grid(row=5, column=0, columnspan=4, pady=20, padx=(110, 0), sticky='ew')

    # Center the input window on the screen
    window_width = 420
    window_height = 250
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2
    input_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    input_window.mainloop()


def main():
    get_user_input()

if __name__ == "__main__":
    main()
